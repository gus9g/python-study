import datetime
import os

# PARA RECEBER PARAMETROS POR INJEÇÃO DE DEPENDENCIAS
# É NECESSARIO QUE SEJA UM METODO/FUNÇÃO
import re

# NECESARIO INSTALAÇÃO DOS PACOTES
from openpyxl import load_workbook, Workbook


# PATH PLANILHA SHARED LIBS
pathSharedLib = "./sharedLibs.xlsx"

if 'sharedLibs.xlsx' not in os.listdir():
    print("A planilha esta sendo Gerada, favor preenche-la")
    print("Certifique-se de manter um backup do seu arquivo classpath")
    wb = Workbook()

    wb.active.title = "nomedoprojeto"
    wb.create_sheet("info")

    libs_sheet = wb["nomedoprojeto"]
    infos_sheet = wb["info"]
    libs_sheet["A1"] = "Libs of Project"
    infos_sheet["A2"] = "Path Project =>"
    infos_sheet["B2"] = "C:\\Examples\\Java\\"
    infos_sheet["A3"] = "Last Line Sheet[0] =>"
    infos_sheet["B3"] = "20"
    wb.save(pathSharedLib)

else:
    print("Iniciando inclusão no Classpath...")

    # CARREGAMENTO DA PLANILHA SHARED LIBS
    fileExcel = load_workbook(pathSharedLib)
    libs_sheet = fileExcel.worksheets[0]
    infos_sheet = fileExcel.worksheets[1]
    cpathsProjects = ".classpath"

    # Testes Criação de Backup Classpath
    pathEclipseProject = infos_sheet["B2"].value + str(libs_sheet.title) + "\\"
    os.system("xcopy " + str(pathEclipseProject) + "'\\.classpath' "
              + str(pathEclipseProject) + "'\\.classpath_'" + str(1))

    # LEITURA DO ARQUIVO CLASSPATH ORIGINAL
    file = open(str(pathEclipseProject) + cpathsProjects, "r")

    # ATRIBUIÇÃO DE TAGS XML, NO VETOR DATA
    data = file.readlines()

    # FECHANDO ARQUIVO ORIGINAL DE LEITURA
    file.close()

    # SALVAMENTO DE CLONE TEMPORARIO E
    # INCLUSÃO DOS CAMPOS
    fileTemp = open(str(pathEclipseProject) + cpathsProjects+"-temp", "w+")

    # CERTIFICAÇÃO DO USUARIO TER IMPORTADO AS USER LIBRARIES
    # NO PROJETO ECLIPSE
    print("Antes de prosseguir, importe o User Library gerado...")
    os.system('pause')

    for dt in data:
        # print("{}".format(dt.strip()))
        fileTemp.write("{}".format(dt.strip()))

        # ADAPTANDO OS DADOS RETORNADOS DA PLANILHA -- TESTES
        if re.search('<classpathentry kind="con" path="org.eclipse.jst.j2ee.internal.module.container"/>', dt):

            for i in range(1, infos_sheet["B3"].value):

                # PREENCHIMENTO DE TAG 'classpathentry' COM DADOS DA PLANILHA COM SHARED LIBRARIES
                if libs_sheet['A'+str(i)].value != None:
                    tagWithUserLibrary = '<classpathentry kind="con" path="org.eclipse.jdt.USER_LIBRARY/' + str(libs_sheet['A'+str(i)].value).upper() + '"/>'

                    # FOR PARA INCLUSÃO DAS USER LIBRARIES ESPECIFICADAS NA PLANILHA
                    data.insert(data.index(dt)+1, tagWithUserLibrary)

    fileTemp.close()


    # ----------------------------------------------------
    # ORGANIZAÇÃO DO CLASSPATH-TEMP PARA O ORIGINAL
    # ----------------------------------------------------

    # LEITURA DO ARQUIVO TEMPORARIO
    fileTemp = open(str(pathEclipseProject) + cpathsProjects + "-temp", "r")

    # -------------------------
    # AUTO INDENTANDO O XML
    # -------------------------

    # ABERTURA DE ARQUIVO PARA ESCRITA/BINARIA (XML)
    filePretty = open(str(pathEclipseProject) + cpathsProjects, "wb+")

    # IMPORTAÇÃO DE BIBLIOTECA PARA AUTO IDENTAÇÃO DE XML
    import lxml.etree as etree

    # SELECIONANDO ARQUIVO XML A SER INDENTADO
    x = etree.parse(fileTemp)

    # AUTO INDENTANDO O XML
    prettyXml = etree.tostring(x, pretty_print=True)

    # SALVANDO A AUTO INDENTAÇÃO EM NOVO XML
    filePretty.write(prettyXml)

    # FECHAMENTO DE ARQUIVOS
    file.close()
    fileTemp.close()
    filePretty.close()

    # EXCLUSÃO DO ARQUIVO TEMPORARIO
    os.remove(str(pathEclipseProject) + cpathsProjects + "-temp")

