# IMPORTAÇÃO DE COMANDOS DO TERMINAL
import os
import time, shutil
from shutil import copyfile

# USO DE CLASSE PESSOAL
from filenames import FilenameDir

try:
    # USO DE PLANILHAS
    from openpyxl import Workbook
    from openpyxl import load_workbook
except:
    os.system("pip install openpyxl")
    print("Bibliotecas instaladas - Execute novamente")

path = "path_and_files.xlsx"

option = int(input("Informe 0 para recriar a planilha\n1 para copia-los\n2 para move-los..."))

if option == 0:
    try:
        #REMOVENDO PLANILHA
        os.remove(path)
    except(FileNotFoundError):
        pass
    
    #CARREGAR PLANILHA EXISTENTE
    file_excel = Workbook()

    #COMANDO UTILIZAR PLANILHA PADRAO
    plan1 = file_excel.active

    plan1['A1'] = 'Nome'
    plan1['B1'] = 'Alterar'

    #SALVAR PLANILHA
    file_excel.save(path)
    
    print("Sua planilha foi recriada!")

if option == 1:
    #CARREGAR PLANILHA EXISTENTE
    file_excel = load_workbook(path)

    #COMANDO UTILIZAR PLANILHA PADRAO
    plan1 = file_excel.active

    lastElement = int(input("Informe a ultima linha preenchida na planilha 'path_and_files.xlsx'...")) + 1

    # FUNÇÃO PARA COPIAR OS ARQUIVOS
    for item in range(2, lastElement):
        shutil.copy(plan1['A'+str(item)].value, plan1['B'+str(item)].value)

    #SALVAR PLANILHA
    file_excel.save(path)

    print("Copia de arquivos concluida com sucesso!")

if option == 2:
    #CARREGAR PLANILHA EXISTENTE
    file_excel = load_workbook(path)

    #COMANDO UTILIZAR PLANILHA PADRAO
    plan1 = file_excel.active

    lastElement = int(input("Informe a ultima linha preenchida na planilha 'path_and_files.xlsx'...")) + 1

    # FUNÇÃO PARA COPIAR OS ARQUIVOS
    for item in range(2, lastElement):
        try:
            shutil.move(plan1['A'+str(item)].value, plan1['B'+str(item)].value)
        except shutil.Error:
            print("O arquivo " + str(plan1['A'+str(item)].value) + ", não pode ser movido para o diretorio " + plan1['B'+str(item)].value)
        
    #SALVAR PLANILHA
    file_excel.save(path)

    print("Movimentação de arquivos concluida!")