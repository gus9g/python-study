import json
from openpyxl import load_workbook, Workbook

option = int(input("Informe: \n1 - Gerar planilha EXCEL atraves de um JSON \n2 - Gerar JSON atraves de um planilha EXCEL"))

# Gerar planilha EXCEL atraves de um JSON
if option == 1:
    lista1 = []
    lista2 = []
    lista3 = []
    lista4 = []
    lista5 = []
    

    a = "cover"
    b = "name"
    c = "url"
    # d = "nameCover"

    print()
    with open('e1.json') as json_file:
        data = json.load(json_file)
        for p in data["teste"]:
            lista1.append(p[a])
            lista2.append(p[b])
            lista3.append(p[c])
            # lista4.append(p[d])

    arquivo_excel = Workbook()
    plan1 = arquivo_excel.active
    
    plan1['a2'] = a
    plan1['b2'] = b
    plan1['c2'] = c
    # plan1['d2'] = d


    for i in range(len(lista1)):
        plan1['a'+str(i+3)] = lista1[i]

    for i in range(len(lista2)):
        plan1['b'+str(i+3)] = lista2[i]
        
    for i in range(len(lista3)):
        plan1['c'+str(i+3)] = lista3[i]
        
    # for i in range(len(lista4)):
    #     plan1['d'+str(i+3)] = lista4[i]


        # print(str(lista1[i]) + " | " + str(lista2[i]) + " | " + str(lista3[i]) + " | " + str(lista4[i]) )

    arquivo_excel.save("exemplo.xlsx")
    print("Planilha atualizada com Sucesso!")

# Gerar JSON atraves de um planilha EXCEL
elif option == 2:
    print("\nAtenção a linha 3 deve conter o nome do campo do JSON")
    print("\nCaso queira dar um nome ao objeto JSON utilize a Celula A1")

    lista = []
    path = "json-create.xlsx"

    nameSheet = "anime"

    amount = int(input("Informe a ultima linha preenchida"))

    arquivo_excel = load_workbook(path)
    
    # plan1 = arquivo_excel.active
    plan1 = arquivo_excel.worksheets[0]

    with open('e1.json','w') as fileJson:
        fileJson.write('{')
        fileJson.write('"')
        fileJson.write(nameSheet)

        fileJson.write('":[')
        for i in range(amount):
            # Incluindo o primeiro campo
            fileJson.write('{')
            fileJson.write('"')
            fileJson.write(str(plan1['A2'].value))
            fileJson.write('": "')
            if str(plan1['A'+str(i+3)].value) != "None":
                fileJson.write(str(plan1['A'+str(i+3)].value))
            else: 
                fileJson.write("")
            fileJson.write('",')

            # Incluindo o primeiro campo
            fileJson.write('"')
            fileJson.write(str(plan1['B2'].value))
            fileJson.write('": "')
            if str(plan1['B'+str(i+3)].value) != "None":
                fileJson.write(str(plan1['B'+str(i+3)].value))
            else:
                fileJson.write("")
            fileJson.write('",')

            # Incluindo o primeiro campo
            fileJson.write('"')
            fileJson.write(str(plan1['C2'].value))
            fileJson.write('": "')
            if str(plan1['C'+str(i+3)].value) != "None":
                fileJson.write(str(plan1['C'+str(i+3)].value))
            else: 
                fileJson.write("")
            fileJson.write('",')


            # Incluindo o primeiro campo
            fileJson.write('"')
            fileJson.write(str(plan1['D2'].value))
            fileJson.write('": "')
            if str(plan1['D'+str(i+3)].value) != "None":
                fileJson.write(str(plan1['D'+str(i+3)].value))
            else: 
                fileJson.write("")
            fileJson.write('"')

            # # Incluindo o primeiro campo
            # fileJson.write('"')
            # fileJson.write(str(plan1['E2'].value))
            # fileJson.write('": "')
            # if str(plan1['E'+str(i+3)].value) != "None":
            #     fileJson.write(str(plan1['E'+str(i+3)].value))
            # else: 
            #     fileJson.write("")
                
            # fileJson.write('",')

            # # Incluindo o primeiro campo
            # fileJson.write('"')
            # fileJson.write(str(plan1['F2'].value))
            # fileJson.write('": "')
            # if str(plan1['F'+str(i+3)].value) != "None":
            #     fileJson.write(str(plan1['F'+str(i+3)].value))
            # else: 
            #     fileJson.write("")
            # fileJson.write('"')


            if(i != amount - 1):
                fileJson.write('},')
            else:
                fileJson.write('}')

            # print(str(plan1['A'+str(i+3)].value))
        fileJson.write(']')
        fileJson.write('}')
        fileJson.close
    arquivo_excel.close
    print("JSON criado com Sucesso!")


# Opção invalida
else:
    print("Opção Invalida, Tente novamente")
    