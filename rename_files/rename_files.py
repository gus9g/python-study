# IMPORTAÇÃO DE COMANDOS DO TERMINAL
import os
import time

# USO DE CLASSE PESSOAL
from filenames import FilenameDir

# Uso de planilhas
from openpyxl import Workbook
from openpyxl import load_workbook


import subprocess


path = "rename_file.xlsx"

option = int(input("Informe 0 para listar arquivos em Excel\n ou 1 para renomea-los..."))

if option == 0:
    try:
        #REMOVENDO PLANILHA
        os.remove(path)
    except(FileNotFoundError):
        pass
    
    #Carregar Planilha Existente
    file_excel = Workbook()

    #COMANDO UTILIZAR PLANILHA PADRAO
    plan1 = file_excel.active
    
    # LISTAR ARQUIVOS DA PASTA EXTERNOS
    fn = FilenameDir("externos")

    plan1['A1'] = 'Nome'
    plan1['B1'] = 'Alterar'
    
    lastElement = 2+len(fn.getName())
    for item in range(2, lastElement):
        plan1['A'+str(item)] = fn.getName()[item - 2]

    #SALVAR PLANILHA
    file_excel.save(path)
    
    print("Sua planilha foi recriada!")

if option == 1:
    #CARREGAR PLANILHA EXISTENTE
    file_excel = load_workbook(path)

    #COMANDO UTILIZAR PLANILHA PADRAO
    plan1 = file_excel["Sheet"]

    lastElement = int(input("Informe a ultima linha preenchida no excel 'rename_file.xlsx'...")) + 1

    # FUNÇÃO PARA RENOMEAR OS ARQUIVOS
    for item in range(2, lastElement):
        # COPIA DOS ARQUIVOS COM CRIAÇÃO DOS DIRETORIOS INEXISTENTES
        subprocess.call(["xcopy", "externos/" + str(plan1['A'+str(item)].value), "externos/" + str(plan1['B'+str(item)].value)])

    #Salvar Planilha
    file_excel.save("rename_file.xlsx")

    print("Renomeacao concluida com sucesso!")